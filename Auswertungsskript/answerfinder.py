import os
import json
from collections import defaultdict

try:
    # Used for writing results into the existing Excel template Mappe1.xlsx
    # Install via "pip install openpyxl" if missing.
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

with open("jsonfile.txt", "r", encoding="utf-8") as f:
    submissions = json.load(f)

# --- average duration ---
durations = [s["durationSeconds"] for s in submissions if "durationSeconds" in s]
avg_duration = sum(durations) / len(durations) if durations else 0.0

# --- counts per pairId (or fallback to comparison index) ---
# answers:
#   0 = "Beide klingen gleich gut"
#   1 = "Audio 1 klingt besser"
#   2 = "Audio 2 klingt besser"
pair_counts = defaultdict(lambda: {0: 0, 1: 0, 2: 0})

for submission in submissions:
    for ans in submission.get("answers", []):
        # Prefer the new stable identifier
        pair_id = ans.get("pairId")

        # Fallback for older data that has only 'comparison'
        if not pair_id:
            comp_idx = ans.get("comparison")
            pair_id = f"comparison_{comp_idx}"

        choice = ans.get("answer")
        if choice in (0, 1, 2):
            pair_counts[pair_id][choice] += 1

# --- console output + prepare data for Excel ---
print(f"Average duration: {avg_duration:.2f} seconds\n")

# store raw counts per pair_id so we can map them into the template later
pair_answer_counts = {}

# Desired ordering of pairs: starting from the original (orig_vs_orig), then
# descending by bitrate from highest (320) to lowest (32), grouped by song.
# This order is used both for console output and Excel export.
PAIR_ORDER = [
    # Bohemian Rhapsody: orig, 320, 224, 128, 64, 32
    "bohemian_orig_vs_orig",
    "bohemian_320_vs_orig",
    "bohemian_224_vs_orig",
    "bohemian_128_vs_orig",
    "bohemian_orig_vs_64",
    "bohemian_orig_vs_32",
    # Detective Conan Theme: orig, 320, 224, 128, 64, 32
    "conan_orig_vs_orig",
    "conan_320_vs_orig",
    "conan_224_vs_orig",
    "conan_orig_vs_128",
    "conan_orig_vs_64",
    "conan_32_vs_orig",
    # Tom's Diner: orig, 320, 224, 128, 64, 32
    "tomsdiner_orig_vs_orig",
    "tomsdiner_320_vs_orig",
    "tomsdiner_orig_vs_224",
    "tomsdiner_128_vs_orig",
    "tomsdiner_64_vs_orig",
    "tomsdiner_orig_vs_32",
]

# For each pair, describe which channel (A1/A2) is original and which bitrate.
# Values are short labels, e.g. "orig" or "320k".
PAIR_AUDIO_INFO = {
    # Bohemian Rhapsody
    "bohemian_orig_vs_orig": ("orig", "orig"),
    "bohemian_320_vs_orig":  ("320k", "orig"),
    "bohemian_224_vs_orig":  ("224k", "orig"),
    "bohemian_128_vs_orig":  ("128k", "orig"),
    "bohemian_orig_vs_64":   ("orig", "64k"),
    "bohemian_orig_vs_32":   ("orig", "32k"),
    # Detective Conan Theme
    "conan_orig_vs_orig":    ("orig", "orig"),
    "conan_320_vs_orig":     ("320k", "orig"),
    "conan_224_vs_orig":     ("224k", "orig"),
    "conan_orig_vs_128":     ("orig", "128k"),
    "conan_orig_vs_64":      ("orig", "64k"),
    "conan_32_vs_orig":      ("32k", "orig"),
    # Tom's Diner
    "tomsdiner_orig_vs_orig": ("orig", "orig"),
    "tomsdiner_320_vs_orig":  ("320k", "orig"),
    "tomsdiner_orig_vs_224":  ("orig", "224k"),
    "tomsdiner_128_vs_orig":  ("128k", "orig"),
    "tomsdiner_64_vs_orig":   ("64k", "orig"),
    "tomsdiner_orig_vs_32":   ("orig", "32k"),
}

# Pretty table header for per-pair summary
print(
    f"{'Pair ID':<28}  "
    f"{'v1':<6}{'v2':<6}"
    f"{'n':>5} | "
    f"{'a0':>5} {'%':>6} | "
    f"{'a1':>5} {'%':>6} | "
    f"{'a2':>5} {'%':>6}"
)
print(
    f"{'-' * 28}  "
    f"{'-' * 6}{'-' * 6}"
    f"{'-' * 5}-+-"
    f"{'-' * 5}-{'-' * 6}-+-"
    f"{'-' * 5}-{'-' * 6}-+-"
    f"{'-' * 5}-{'-' * 6}"
)

# Use PAIR_ORDER for console output; append any unexpected pairIds at the end.
ordered_for_console = [pid for pid in PAIR_ORDER if pid in pair_counts]
remaining_ids = sorted(pid for pid in pair_counts.keys() if pid not in ordered_for_console)
all_console_ids = ordered_for_console + remaining_ids

for idx, pair_id in enumerate(all_console_ids):
    # Insert horizontal dashed separators between songs (after each block of 6 pairs)
    if idx in (6, 12):
        print(
            f"{'-' * 28}  "
            f"{'-' * 6}{'-' * 6}"
            f"{'-' * 5}-+-"
            f"{'-' * 5}-{'-' * 6}-+-"
            f"{'-' * 5}-{'-' * 6}-+-"
            f"{'-' * 5}-{'-' * 6}"
        )

    counts = pair_counts[pair_id]
    total = sum(counts.values())
    if total == 0:
        continue

    p0 = counts[0] / total * 100
    p1 = counts[1] / total * 100
    p2 = counts[2] / total * 100

    # table row: pair name, A1/A2 info, n, counts + percentages
    a1_info, a2_info = PAIR_AUDIO_INFO.get(pair_id, ("?", "?"))
    print(
        f"{pair_id:<28}  "
        f"{a1_info:<6}{a2_info:<6}"
        f"{total:>5} | "
        f"{counts[0]:>5} {p0:>5.1f}% | "
        f"{counts[1]:>5} {p1:>5.1f}% | "
        f"{counts[2]:>5} {p2:>5.1f}%"
    )

    # store raw counts; columns B/C/D in the template will hold counts for answers 0/1/2
    pair_answer_counts[pair_id] = (counts[0], counts[1], counts[2])

# --- write into existing Excel template Mappe1.xlsx ---
if load_workbook is None:
    print("\n[INFO] openpyxl is not installed; skipping Excel export. "
          "Install it with 'pip install openpyxl' and run this script again.")
else:
    # Resolve Mappe1.xlsx relative to this script file, not the current working directory.
    # Mappe1.xlsx currently lives in the same folder as this script (extras/).
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "Mappe1.xlsx")
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        print(f"\n[WARN] Could not find Excel template at '{template_path}'. "
              "Skipping Excel export.")
    else:
        ws = wb.active  # use the first sheet; adjust if you need a specific sheet name

        # Take all relevant pairIds, ordered by ascending bitrate as defined in PAIR_ORDER,
        # and write their COUNTS into the correct per-song regions of the template:
        #   Bohemian Rhapsody: columns B/C/D, rows 2-7  (6 Vergleiche)
        #   Detective Conan:   columns G/H/I, rows 2-7
        #   Tom's Diner:       columns L/M/N, rows 2-7
        ordered_ids = [pid for pid in PAIR_ORDER if pid in pair_answer_counts]

        for idx, pair_id in enumerate(ordered_ids):
            # Determine which song this pair belongs to and which row inside that block.
            song_index = idx // 6      # 0 = Bohemian, 1 = Conan, 2 = Tom's Diner
            within_song = idx % 6      # 0..5 -> rows 2..7

            if song_index > 2:
                # Safety: ignore anything beyond the three defined songs.
                break

            row = 2 + within_song

            if song_index == 0:
                base_col = 2   # B/C/D for Bohemian Rhapsody
            elif song_index == 1:
                base_col = 7   # G/H/I for Detective Conan
            else:
                base_col = 12  # L/M/N for Tom's Diner

            c0, c1, c2 = pair_answer_counts[pair_id]
            ws.cell(row=row, column=base_col,     value=c0)
            ws.cell(row=row, column=base_col + 1, value=c1)
            ws.cell(row=row, column=base_col + 2, value=c2)

        # Write n (total answers per pair) into S2. Since each synthetic dataset
        # uses every pairId equally often, we can take n from any one of the
        # ordered_ids if available.
        if ordered_ids:
            first_id = ordered_ids[0]
            n0, n1, n2 = pair_answer_counts[first_id]
            ws.cell(row=2, column=19, value=(n0 + n1 + n2))  # column 19 == 'S'

        output_filename = "Mappe1_filled.xlsx"
        wb.save(output_filename)
        print(f"\nExcel export written to '{output_filename}'.")
